"""
Parser for the Commander's historical xlsx file.

Tab: 'import list'
Layout:
  Row 1  — title row (ignored)
  Row 2  — column headers (16 columns; see COLUMN_MAP below)
  Row 3+ — data rows (576 transactions)

Actual column headers (row 2, confirmed 2026-05-28):
  Year | Month | Day | Importer | Translation of type | Country of origin |
  Country of export | Importer (Korean) | Product name (Korean) |
  Product name (English) | Product type (Korean) | Exporter (English) |
  Date | Expire date start from | Country of origin (KR) | Country of export (KR)

Output schema (internal English field names, matching Sheets tab headers):
  year, month, day, date, importer_en, importer_ko,
  product_ko, product_en, product_type_ko, product_type_en,
  country_origin_en, country_origin_ko, country_export_en, country_export_ko,
  exporter_en, expiry_date,
  type_frozen, type_dried, type_ambiguous
"""

from __future__ import annotations

import datetime
from pathlib import Path
from typing import Any

import openpyxl

from config import cfg

_IMPORT_LIST_TAB = "import list"

# Map from raw xlsx header → internal field name
COLUMN_MAP: dict[str, str] = {
    "Year": "year",
    "Month": "month",
    "Day": "day",
    "Importer": "importer_en",
    "Translation of type": "product_type_en",
    "Country of origin": "country_origin_en",
    "Country of export": "country_export_en",
    "Importer (Korean)": "importer_ko",
    "Product name (Korean)": "product_ko",
    "Product name (English)": "product_en",
    "Product type (Korean)": "product_type_ko",
    "Exporter (English)": "exporter_en",
    "Date": "date_raw",
    "Expire date start from": "expiry_date_raw",
    "Country of origin (KR)": "country_origin_ko",
    "Country of export (KR)": "country_export_ko",
}

# Product types that are clearly frozen
_FROZEN_KEYWORDS = frozenset(["냉동", "생녹용"])
# Product types that are clearly dried
_DRIED_KEYWORDS = frozenset(["건조", "건"])


def _classify_type(product_type_ko: str, product_type_en: str = "") -> dict[str, bool]:
    """Return frozen/dried/ambiguous flags.

    Uses Korean product type as primary signal; falls back to English
    translation when Korean is empty (320 historical rows have no KO value).
    """
    kw_ko = str(product_type_ko).strip()
    kw_en = str(product_type_en).strip().lower()

    # Korean-based classification
    if kw_ko:
        is_frozen = any(k in kw_ko for k in _FROZEN_KEYWORDS)
        is_dried = "건조" in kw_ko or (kw_ko.startswith("건") and "건강" not in kw_ko)
    else:
        # Fall back to English translation column
        is_frozen = "frozen" in kw_en
        is_dried = "dried" in kw_en

    if is_frozen and is_dried:
        return {"type_frozen": False, "type_dried": False, "type_ambiguous": True}
    if is_frozen:
        return {"type_frozen": True, "type_dried": False, "type_ambiguous": False}
    if is_dried:
        return {"type_frozen": False, "type_dried": True, "type_ambiguous": False}
    # e.g. plain 녹용 or Hard antlers — ambiguous
    return {"type_frozen": False, "type_dried": False, "type_ambiguous": True}


def _parse_date(value: Any) -> str:
    """Return an ISO-format date string (YYYY-MM-DD) or empty string."""
    if value is None:
        return ""
    if isinstance(value, (datetime.date, datetime.datetime)):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if not s or s == "-":
        return ""
    return s


def _safe_str(value: Any) -> str:
    """Return stripped string, or empty string for None."""
    if value is None:
        return ""
    return str(value).strip()


def parse(xlsx_path: Path | None = None) -> list[dict[str, Any]]:
    """Parse the 'import list' tab and return a list of normalised row dicts.

    Args:
        xlsx_path: Path to the xlsx file. Defaults to cfg.historical_xlsx_path.

    Returns:
        List of dicts with internal English field names. One dict per data row.
        Blank rows and the header rows are skipped automatically.

    Raises:
        ValueError: If the expected tab or expected headers are not found.
    """
    path = xlsx_path or cfg.historical_xlsx_path

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if _IMPORT_LIST_TAB not in wb.sheetnames:
        raise ValueError(
            f"Tab '{_IMPORT_LIST_TAB}' not found in {path}. "
            f"Available tabs: {wb.sheetnames}"
        )

    ws = wb[_IMPORT_LIST_TAB]
    all_rows = list(ws.iter_rows(values_only=True))

    # Row 1 (index 0) is the title row — skip it.
    # Row 2 (index 1) is the header row.
    if len(all_rows) < 2:
        raise ValueError(f"'{_IMPORT_LIST_TAB}' tab has fewer than 2 rows — cannot parse.")

    raw_headers = [_safe_str(h) for h in all_rows[1]]

    # Validate that expected columns are present
    missing = [col for col in COLUMN_MAP if col not in raw_headers]
    if missing:
        raise ValueError(
            f"Missing expected columns in '{_IMPORT_LIST_TAB}': {missing}\n"
            f"Actual headers: {raw_headers}"
        )

    # Build index map: column name → position
    col_index: dict[str, int] = {h: i for i, h in enumerate(raw_headers) if h}

    results: list[dict[str, Any]] = []

    for raw_row in all_rows[2:]:  # data starts at row 3 (index 2)
        # Skip completely empty rows
        if all(v is None or str(v).strip() == "" for v in raw_row):
            continue

        row: dict[str, Any] = {}

        # Map raw columns to internal names
        for raw_col, internal_name in COLUMN_MAP.items():
            idx = col_index.get(raw_col)
            if idx is None or idx >= len(raw_row):
                row[internal_name] = ""
            else:
                row[internal_name] = raw_row[idx]

        # Normalise date fields
        row["date"] = _parse_date(row.pop("date_raw", ""))
        row["expiry_date"] = _parse_date(row.pop("expiry_date_raw", ""))

        # Normalise string fields
        for field in [
            "importer_en", "importer_ko", "product_ko", "product_en",
            "product_type_ko", "product_type_en", "exporter_en",
            "country_origin_en", "country_origin_ko",
            "country_export_en", "country_export_ko",
        ]:
            row[field] = _safe_str(row.get(field, ""))

        # Normalise year/month/day
        row["year"] = _safe_str(row.get("year", ""))
        row["month"] = _safe_str(row.get("month", ""))
        row["day"] = _safe_str(row.get("day", ""))

        # Add frozen/dried classification flags
        # Pass both KO and EN: KO is primary; EN is fallback for 320 rows where KO is empty
        row.update(_classify_type(
            row.get("product_type_ko", ""),
            row.get("product_type_en", ""),
        ))

        # Source marker
        row["source"] = "historical"

        results.append(row)

    return results


def get_expected_headers() -> list[str]:
    """Return the full list of internal field names in canonical order.

    Used by the Sheets client to create the tab header row.
    """
    return [
        "year", "month", "day", "date",
        "importer_en", "importer_ko",
        "product_ko", "product_en",
        "product_type_ko", "product_type_en",
        "country_origin_en", "country_origin_ko",
        "country_export_en", "country_export_ko",
        "exporter_en", "expiry_date",
        "type_frozen", "type_dried", "type_ambiguous",
        "source",
    ]
