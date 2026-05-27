"""
Translation table loader for Velvet Food Imports.

Reads the 'Names for Vlookup' tab from the Commander's historical xlsx
and returns lookup dicts for four categories:
  - countries_of_origin  (Korean → English)
  - product_types        (Korean → English)
  - importers            (Korean → English)
  - exporters            (English company name → country string)

Design rule: unknown keys ALWAYS return the original string — never raise.
This ensures new entries in MFDS files do not break the pipeline.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

from config import cfg

# Tab name in the Commander xlsx
_VLOOKUP_TAB = "Names for Vlookup"


def _load_raw_rows(xlsx_path: Path) -> list[tuple]:
    """Load all non-empty rows from the Vlookup tab as raw tuples."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if _VLOOKUP_TAB not in wb.sheetnames:
        raise ValueError(f"Tab '{_VLOOKUP_TAB}' not found in {xlsx_path}")
    ws = wb[_VLOOKUP_TAB]
    rows = []
    for row in ws.iter_rows(values_only=True):
        # Each row is (None, col_b, col_c) — we only care about B and C
        if len(row) >= 3 and row[1] is not None and str(row[1]).strip():
            rows.append(row)
    return rows


def _safe_str(value: Any) -> str:
    """Return stripped string or empty string for None."""
    if value is None:
        return ""
    return str(value).strip()


class TranslationTable:
    """Holds all four KO→EN lookup dicts loaded from the Vlookup tab.

    Attributes:
        countries_of_origin: Korean country name → English country name.
        product_types:       Korean product type → English product type.
        importers:           Korean company name → English company name.
        exporters:           English exporter name → country of origin string.
    """

    def __init__(self, xlsx_path: Path | None = None) -> None:
        path = xlsx_path or cfg.historical_xlsx_path
        rows = _load_raw_rows(path)

        self.countries_of_origin: dict[str, str] = {}
        self.product_types: dict[str, str] = {}
        self.importers: dict[str, str] = {}
        self.exporters: dict[str, str] = {}

        self._parse(rows)

    def _parse(self, rows: list[tuple]) -> None:
        """Detect section headers and populate each lookup dict."""
        # Sections are separated by sentinel rows that contain known labels.
        # We detect the current section by the content of the sentinel row.
        mode: str | None = None

        for row in rows:
            b = _safe_str(row[1])
            c = _safe_str(row[2]) if len(row) > 2 else ""

            # Detect section headers
            if b == "뉴질랜드" and c == "New Zealand":
                mode = "country_of_origin"
                self.countries_of_origin[b] = c
                continue
            if b == "Product name (Korean)" and c == "Translation":
                mode = "product_type"
                continue
            if b == "Company Korean name" and c == "Company English name":
                mode = "importer"
                continue
            if b == "Company" and c == "Country":
                mode = "exporter"
                continue

            # Skip rows without a value in both B and C
            if not b or not c:
                continue

            # Populate the active section dict
            if mode == "country_of_origin":
                self.countries_of_origin[b] = c
            elif mode == "product_type":
                self.product_types[b] = c
            elif mode == "importer":
                self.importers[b] = c
            elif mode == "exporter":
                self.exporters[b] = c

    # ------------------------------------------------------------------ #
    # Lookup helpers — unknown keys return original string, never raise   #
    # ------------------------------------------------------------------ #

    def country_of_origin(self, korean: str) -> str:
        """Map a Korean country name to English. Falls back to original."""
        return self.countries_of_origin.get(korean.strip(), korean.strip())

    def product_type(self, korean: str) -> str:
        """Map a Korean product type to English. Falls back to original."""
        return self.product_types.get(korean.strip(), korean.strip())

    def importer(self, korean: str) -> str:
        """Map a Korean importer name to English. Falls back to original."""
        return self.importers.get(korean.strip(), korean.strip())

    def exporter_country(self, english_exporter: str) -> str:
        """Return the country for a known English exporter. Falls back to original."""
        return self.exporters.get(english_exporter.strip(), english_exporter.strip())

    def is_frozen(self, product_ko: str) -> bool:
        """Return True if the Korean product name indicates frozen (냉동) velvet."""
        kw = product_ko.strip()
        return "냉동" in kw or "생녹용" in kw

    def is_dried(self, product_ko: str) -> bool:
        """Return True if the Korean product name indicates dried (건조) velvet."""
        kw = product_ko.strip()
        return "건조" in kw or ("냉동" not in kw and "생" not in kw and "녹용" in kw)


# Module-level singleton — lazily loaded from default path
_table: TranslationTable | None = None


def get_table(xlsx_path: Path | None = None) -> TranslationTable:
    """Return (and cache) the global TranslationTable instance.

    Pass xlsx_path explicitly in tests; omit in production to use cfg default.
    """
    global _table  # noqa: PLW0603
    if xlsx_path is not None:
        return TranslationTable(xlsx_path)
    if _table is None:
        _table = TranslationTable()
    return _table
