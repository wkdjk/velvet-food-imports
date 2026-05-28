"""
Unit tests for src/parsers/historical_xlsx.py and src/parsers/mfds_xlsx.py  (A-24).

All fixtures are built entirely in memory (openpyxl Workbook → temp file).
No real Commander xlsx file, no Google Sheets, no credentials required.

Run:  PYTHONPATH=. python -m pytest tests/test_parsers.py -v
"""

from __future__ import annotations

import datetime
from pathlib import Path

import openpyxl
import pytest

# ---------------------------------------------------------------------------
# Helpers — synthetic xlsx builders
# ---------------------------------------------------------------------------

# These are the 16 headers as they appear in row 2 of the 'import list' tab.
_HIST_HEADERS = [
    "Year", "Month", "Day", "Importer", "Translation of type",
    "Country of origin", "Country of export", "Importer (Korean)",
    "Product name (Korean)", "Product name (English)",
    "Product type (Korean)", "Exporter (English)", "Date",
    "Expire date start from", "Country of origin (KR)", "Country of export (KR)",
]

# These are the 14 Korean headers for MFDS files.
_MFDS_HEADERS = [
    "처리일자", "수입업체", "신고번호", "품목제조번호",
    "제품명(한글)", "제품명(영문)", "품목류", "수출국",
    "제조국", "제조사(영문)", "순중량(KG)", "수량(갯수)",
    "유통기한", "원산지",
]

# A minimal 'Names for Vlookup' tab so the MFDS parser's translation table loads.
def _add_vlookup_tab(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Names for Vlookup")
    ws.append([None, "뉴질랜드", "New Zealand"])
    ws.append([None, "러시아",   "Russia"])
    ws.append([None, "Product name (Korean)", "Translation"])
    ws.append([None, "생녹용(냉동)", "Frozen velvet antler"])
    ws.append([None, "녹용(건조)",   "Dried velvet antler"])
    ws.append([None, "Company Korean name", "Company English name"])
    ws.append([None, "한국수입사", "Korea Importer Co"])
    ws.append([None, "Company", "Country"])
    ws.append([None, "NZ Velvet Exports", "New Zealand"])


def _hist_xlsx(tmp_path: Path, data_rows: list[list] | None = None) -> Path:
    """Build a minimal historical xlsx with an 'import list' tab.

    Row 1 = title row (ignored by parser).
    Row 2 = column headers.
    Rows 3+ = data rows supplied by caller (default: two sample rows).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "import list"

    # Row 1 — title (ignored)
    ws.append(["Korean Food Imports"] + [""] * 15)

    # Row 2 — headers
    ws.append(_HIST_HEADERS)

    # Data rows
    if data_rows is None:
        # Two canonical rows: one frozen, one dried
        ws.append([
            2020, 1, 15,
            "Korea Importer Co",          # Importer
            "Frozen velvet antler",        # Translation of type
            "New Zealand",                 # Country of origin
            "New Zealand",                 # Country of export
            "한국수입사",                  # Importer (Korean)
            "생녹용(냉동)",                # Product name (Korean)
            "Frozen velvet antler",        # Product name (English)
            "생녹용(냉동)",                # Product type (Korean)
            "NZ Velvet Exports",           # Exporter (English)
            datetime.date(2020, 1, 15),    # Date
            datetime.date(2022, 1, 15),    # Expire date start from
            "뉴질랜드",                    # Country of origin (KR)
            "뉴질랜드",                    # Country of export (KR)
        ])
        ws.append([
            2020, 3, 10,
            "Seoul Velvet Ltd",            # Importer
            "Dried velvet antler",         # Translation of type
            "Russia",                      # Country of origin
            "Russia",                      # Country of export
            "서울벨벳",                    # Importer (Korean)
            "녹용(건조)",                  # Product name (Korean)
            "Dried velvet antler",         # Product name (English)
            "녹용(건조)",                  # Product type (Korean)
            "Siberian Antler",             # Exporter (English)
            datetime.date(2020, 3, 10),    # Date
            None,                          # Expire date — intentionally empty
            "러시아",                      # Country of origin (KR)
            "러시아",                      # Country of export (KR)
        ])
    else:
        for row in data_rows:
            ws.append(row)

    _add_vlookup_tab(wb)

    out = tmp_path / "historical.xlsx"
    wb.save(str(out))
    return out


def _mfds_xlsx(tmp_path: Path, data_rows: list[list] | None = None) -> Path:
    """Build a minimal MFDS xlsx with the 14-column header on row 1."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws.append(_MFDS_HEADERS)

    if data_rows is None:
        ws.append([
            "20260515",          # 처리일자
            "한국수입사",        # 수입업체
            "A12345",            # 신고번호
            "M67890",            # 품목제조번호
            "생녹용(냉동)",      # 제품명(한글)
            "Frozen velvet antler",  # 제품명(영문)
            "생녹용(냉동)",      # 품목류
            "뉴질랜드",          # 수출국
            "뉴질랜드",          # 제조국
            "NZ Velvet Exports", # 제조사(영문)
            "500",               # 순중량(KG)
            "100",               # 수량(갯수)
            "20280515",          # 유통기한
            "뉴질랜드",          # 원산지
        ])
        ws.append([
            "20260520",
            "서울벨벳",
            "B99999",
            "N11111",
            "녹용(건조)",
            "Dried velvet antler",
            "녹용(건조)",
            "러시아",
            "러시아",
            "Siberian Antler",
            "300",
            "50",
            None,                # 유통기한 intentionally missing
            "러시아",
        ])
    else:
        for row in data_rows:
            ws.append(row)

    out = tmp_path / "mfds.xlsx"
    wb.save(str(out))
    return out


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def hist_path(tmp_path_factory: pytest.TempPathFactory) -> Path:
    tmp = tmp_path_factory.mktemp("hist")
    return _hist_xlsx(tmp)


@pytest.fixture(scope="module")
def mfds_path(tmp_path_factory: pytest.TempPathFactory) -> Path:
    tmp = tmp_path_factory.mktemp("mfds")
    return _mfds_xlsx(tmp)


# ---------------------------------------------------------------------------
# historical_xlsx.py tests
# ---------------------------------------------------------------------------

class TestHistoricalColumnMapping:
    """Parser correctly maps all 16 raw headers to internal field names."""

    def test_expected_fields_present(self, hist_path: Path) -> None:
        from src.parsers.historical_xlsx import parse, get_expected_headers
        rows = parse(hist_path)
        expected = set(get_expected_headers())
        actual = set(rows[0].keys())
        assert expected.issubset(actual), f"Missing fields: {expected - actual}"

    def test_importer_en_mapped_correctly(self, hist_path: Path) -> None:
        from src.parsers.historical_xlsx import parse
        rows = parse(hist_path)
        assert rows[0]["importer_en"] == "Korea Importer Co"

    def test_importer_ko_mapped_correctly(self, hist_path: Path) -> None:
        from src.parsers.historical_xlsx import parse
        rows = parse(hist_path)
        assert rows[0]["importer_ko"] == "한국수입사"

    def test_exporter_en_mapped_correctly(self, hist_path: Path) -> None:
        from src.parsers.historical_xlsx import parse
        rows = parse(hist_path)
        assert rows[0]["exporter_en"] == "NZ Velvet Exports"

    def test_product_ko_mapped_correctly(self, hist_path: Path) -> None:
        from src.parsers.historical_xlsx import parse
        rows = parse(hist_path)
        assert rows[0]["product_ko"] == "생녹용(냉동)"

    def test_country_export_ko_mapped_correctly(self, hist_path: Path) -> None:
        from src.parsers.historical_xlsx import parse
        rows = parse(hist_path)
        assert rows[0]["country_export_ko"] == "뉴질랜드"


class TestHistoricalRowCount:
    def test_two_data_rows_parsed(self, hist_path: Path) -> None:
        from src.parsers.historical_xlsx import parse
        assert len(parse(hist_path)) == 2

    def test_blank_rows_skipped(self, tmp_path: Path) -> None:
        """Completely blank rows must not appear in the output."""
        # Insert a blank row between the two data rows
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "import list"
        ws.append(["Title"] + [""] * 15)      # row 1 — title
        ws.append(_HIST_HEADERS)               # row 2 — headers
        ws.append([2020, 1, 1, "A", "Frozen velvet antler", "New Zealand", "New Zealand",
                   "한국수입사", "생녹용(냉동)", "Frozen velvet antler",
                   "생녹용(냉동)", "NZ Velvet Exports",
                   datetime.date(2020, 1, 1), None, "뉴질랜드", "뉴질랜드"])
        ws.append([None] * 16)                 # blank row — must be skipped
        ws.append([2020, 2, 1, "B", "Dried velvet antler", "Russia", "Russia",
                   "서울벨벳", "녹용(건조)", "Dried velvet antler",
                   "녹용(건조)", "Siberian Antler",
                   datetime.date(2020, 2, 1), None, "러시아", "러시아"])
        _add_vlookup_tab(wb)
        p = tmp_path / "blanks.xlsx"
        wb.save(str(p))

        from src.parsers.historical_xlsx import parse
        rows = parse(p)
        assert len(rows) == 2


class TestHistoricalTypeDetection:
    def test_frozen_row_flagged_correctly(self, hist_path: Path) -> None:
        from src.parsers.historical_xlsx import parse
        rows = parse(hist_path)
        frozen_rows = [r for r in rows if r["product_type_ko"] == "생녹용(냉동)"]
        assert frozen_rows, "No frozen row found in fixture"
        r = frozen_rows[0]
        assert r["type_frozen"] is True
        assert r["type_dried"] is False
        assert r["type_ambiguous"] is False

    def test_dried_row_flagged_correctly(self, hist_path: Path) -> None:
        from src.parsers.historical_xlsx import parse
        rows = parse(hist_path)
        dried_rows = [r for r in rows if r["product_type_ko"] == "녹용(건조)"]
        assert dried_rows, "No dried row found in fixture"
        r = dried_rows[0]
        assert r["type_frozen"] is False
        assert r["type_dried"] is True
        assert r["type_ambiguous"] is False

    def test_ambiguous_type_falls_back_to_en(self, tmp_path: Path) -> None:
        """Rows with no Korean product type use the English translation column."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "import list"
        ws.append(["Title"] + [""] * 15)
        ws.append(_HIST_HEADERS)
        # Product type (Korean) is empty; English says 'Frozen velvet antler'
        ws.append([2020, 1, 1, "A", "Frozen velvet antler", "New Zealand", "New Zealand",
                   "한국수입사", "생녹용(냉동)", "Frozen velvet antler",
                   "",                           # Product type (Korean) — empty
                   "NZ Velvet Exports",
                   datetime.date(2020, 1, 1), None, "뉴질랜드", "뉴질랜드"])
        _add_vlookup_tab(wb)
        p = tmp_path / "en_fallback.xlsx"
        wb.save(str(p))

        from src.parsers.historical_xlsx import parse
        rows = parse(p)
        assert len(rows) == 1
        assert rows[0]["type_frozen"] is True


class TestHistoricalDateParsing:
    def test_date_column_is_iso_string(self, hist_path: Path) -> None:
        from src.parsers.historical_xlsx import parse
        rows = parse(hist_path)
        assert rows[0]["date"] == "2020-01-15"

    def test_expiry_date_missing_returns_empty(self, hist_path: Path) -> None:
        """Row 2 in the fixture has None for expiry_date — must return empty string."""
        from src.parsers.historical_xlsx import parse
        rows = parse(hist_path)
        assert rows[1]["expiry_date"] == ""

    def test_expiry_date_present_parsed(self, hist_path: Path) -> None:
        from src.parsers.historical_xlsx import parse
        rows = parse(hist_path)
        assert rows[0]["expiry_date"] == "2022-01-15"


class TestHistoricalEdgeCases:
    def test_missing_required_column_raises(self, tmp_path: Path) -> None:
        """If a required column is absent, ValueError must be raised."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "import list"
        ws.append(["Title"])
        # Drop 'Year' from the header row
        ws.append(_HIST_HEADERS[1:])   # missing 'Year'
        _add_vlookup_tab(wb)
        p = tmp_path / "bad_headers.xlsx"
        wb.save(str(p))

        from src.parsers.historical_xlsx import parse
        with pytest.raises(ValueError, match="Missing expected columns"):
            parse(p)

    def test_wrong_tab_name_raises(self, tmp_path: Path) -> None:
        """If the tab is not named 'import list', ValueError must be raised."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "wrong tab"
        ws.append(_HIST_HEADERS)
        p = tmp_path / "wrong_tab.xlsx"
        wb.save(str(p))

        from src.parsers.historical_xlsx import parse
        with pytest.raises(ValueError, match="import list"):
            parse(p)

    def test_source_field_is_historical(self, hist_path: Path) -> None:
        from src.parsers.historical_xlsx import parse
        rows = parse(hist_path)
        assert all(r["source"] == "historical" for r in rows)


# ---------------------------------------------------------------------------
# mfds_xlsx.py tests
# ---------------------------------------------------------------------------

class TestMfdsColumnMapping:
    """Parser correctly maps all 14 Korean headers to internal field names."""

    def test_all_14_internal_fields_present(self, hist_path: Path, mfds_path: Path) -> None:
        from src.parsers.mfds_xlsx import parse, get_expected_headers
        rows = parse(mfds_path, xlsx_path_for_table=hist_path)
        expected = {
            "date", "importer_ko", "report_no", "item_no",
            "product_ko", "product_en", "product_type_ko",
            "country_export_ko", "country_origin_ko",
            "exporter_en", "weight_kg", "quantity",
            "expiry_date", "country_origin_raw",
        }
        actual = set(rows[0].keys())
        assert expected.issubset(actual), f"Missing fields: {expected - actual}"

    def test_importer_ko_extracted(self, hist_path: Path, mfds_path: Path) -> None:
        from src.parsers.mfds_xlsx import parse
        rows = parse(mfds_path, xlsx_path_for_table=hist_path)
        assert rows[0]["importer_ko"] == "한국수입사"

    def test_report_no_extracted(self, hist_path: Path, mfds_path: Path) -> None:
        from src.parsers.mfds_xlsx import parse
        rows = parse(mfds_path, xlsx_path_for_table=hist_path)
        assert rows[0]["report_no"] == "A12345"

    def test_weight_kg_as_string(self, hist_path: Path, mfds_path: Path) -> None:
        """weight_kg must be stored as a string (Sheets type-coercion rule)."""
        from src.parsers.mfds_xlsx import parse
        rows = parse(mfds_path, xlsx_path_for_table=hist_path)
        assert isinstance(rows[0]["weight_kg"], str)
        assert rows[0]["weight_kg"] == "500"


class TestMfdsKoEnNormalisation:
    def test_importer_en_resolved_via_table(self, hist_path: Path, mfds_path: Path) -> None:
        """Korean importer name must be translated to English via the lookup table."""
        from src.parsers.mfds_xlsx import parse
        rows = parse(mfds_path, xlsx_path_for_table=hist_path)
        assert rows[0]["importer_en"] == "Korea Importer Co"

    def test_country_export_en_resolved(self, hist_path: Path, mfds_path: Path) -> None:
        from src.parsers.mfds_xlsx import parse
        rows = parse(mfds_path, xlsx_path_for_table=hist_path)
        assert rows[0]["country_export_en"] == "New Zealand"

    def test_unknown_importer_falls_back_to_original(
        self, tmp_path: Path, hist_path: Path
    ) -> None:
        """An importer not in the table must return the original Korean string."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(_MFDS_HEADERS)
        ws.append([
            "20260515", "미등록수입사", "X001", "Y001",
            "생녹용(냉동)", "Frozen velvet antler", "생녹용(냉동)",
            "뉴질랜드", "뉴질랜드", "NZ Velvet Exports",
            "100", "20", "20280515", "뉴질랜드",
        ])
        p = tmp_path / "unknown_importer.xlsx"
        wb.save(str(p))

        from src.parsers.mfds_xlsx import parse
        rows = parse(p, xlsx_path_for_table=hist_path)
        assert rows[0]["importer_en"] == "미등록수입사"


class TestMfdsTypeFlag:
    def test_frozen_row_classified(self, hist_path: Path, mfds_path: Path) -> None:
        from src.parsers.mfds_xlsx import parse
        rows = parse(mfds_path, xlsx_path_for_table=hist_path)
        frozen = [r for r in rows if r["product_type_ko"] == "생녹용(냉동)"]
        assert frozen, "No frozen row in MFDS fixture"
        assert frozen[0]["type_frozen"] is True
        assert frozen[0]["type_dried"] is False

    def test_dried_row_classified(self, hist_path: Path, mfds_path: Path) -> None:
        from src.parsers.mfds_xlsx import parse
        rows = parse(mfds_path, xlsx_path_for_table=hist_path)
        dried = [r for r in rows if r["product_type_ko"] == "녹용(건조)"]
        assert dried, "No dried row in MFDS fixture"
        assert dried[0]["type_dried"] is True
        assert dried[0]["type_frozen"] is False


class TestMfdsDateParsing:
    def test_yyyymmdd_compact_format_parsed(self, hist_path: Path, mfds_path: Path) -> None:
        """MFDS compact YYYYMMDD date string must be converted to ISO YYYY-MM-DD."""
        from src.parsers.mfds_xlsx import parse
        rows = parse(mfds_path, xlsx_path_for_table=hist_path)
        assert rows[0]["date"] == "2026-05-15"

    def test_year_month_day_derived_from_date(self, hist_path: Path, mfds_path: Path) -> None:
        from src.parsers.mfds_xlsx import parse
        rows = parse(mfds_path, xlsx_path_for_table=hist_path)
        assert rows[0]["year"] == "2026"
        assert rows[0]["month"] == "05"
        assert rows[0]["day"] == "15"

    def test_missing_expiry_date_returns_empty(self, hist_path: Path, mfds_path: Path) -> None:
        """Row 2 in the MFDS fixture has no expiry date — must return empty string."""
        from src.parsers.mfds_xlsx import parse
        rows = parse(mfds_path, xlsx_path_for_table=hist_path)
        assert rows[1]["expiry_date"] == ""


class TestMfdsEdgeCases:
    def test_source_field_is_mfds(self, hist_path: Path, mfds_path: Path) -> None:
        from src.parsers.mfds_xlsx import parse
        rows = parse(mfds_path, xlsx_path_for_table=hist_path)
        assert all(r["source"] == "mfds" for r in rows)

    def test_missing_mfds_columns_raises(self, tmp_path: Path, hist_path: Path) -> None:
        """If a required MFDS column is missing, ValueError must be raised."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        # Drop '처리일자' (and several others) so that the header check fails
        ws.append(_MFDS_HEADERS[3:])   # truncated — missing first 3 required columns
        p = tmp_path / "bad_mfds.xlsx"
        wb.save(str(p))

        from src.parsers.mfds_xlsx import parse
        with pytest.raises(ValueError):
            parse(p, xlsx_path_for_table=hist_path)

    def test_title_row_before_headers_handled(
        self, tmp_path: Path, hist_path: Path
    ) -> None:
        """Some MFDS files have a title row above the column headers; parser must cope."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["수입식품 조회 결과"] + [""] * 13)  # title row
        ws.append(_MFDS_HEADERS)                        # real header on row 2
        ws.append([
            "20260515", "한국수입사", "A12345", "M67890",
            "생녹용(냉동)", "Frozen velvet antler", "생녹용(냉동)",
            "뉴질랜드", "뉴질랜드", "NZ Velvet Exports",
            "500", "100", "20280515", "뉴질랜드",
        ])
        p = tmp_path / "title_row.xlsx"
        wb.save(str(p))

        from src.parsers.mfds_xlsx import parse
        rows = parse(p, xlsx_path_for_table=hist_path)
        assert len(rows) == 1
        assert rows[0]["date"] == "2026-05-15"

    def test_blank_rows_skipped(self, tmp_path: Path, hist_path: Path) -> None:
        """Blank rows between data rows must be silently skipped."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(_MFDS_HEADERS)
        ws.append([
            "20260515", "한국수입사", "A1", "M1",
            "생녹용(냉동)", "Frozen velvet antler", "생녹용(냉동)",
            "뉴질랜드", "뉴질랜드", "NZ Velvet Exports",
            "500", "100", "20280515", "뉴질랜드",
        ])
        ws.append([None] * 14)   # blank row
        ws.append([
            "20260520", "서울벨벳", "B1", "N1",
            "녹용(건조)", "Dried velvet antler", "녹용(건조)",
            "러시아", "러시아", "Siberian Antler",
            "300", "50", None, "러시아",
        ])
        p = tmp_path / "mfds_blanks.xlsx"
        wb.save(str(p))

        from src.parsers.mfds_xlsx import parse
        rows = parse(p, xlsx_path_for_table=hist_path)
        assert len(rows) == 2

    def test_bilingual_product_name_preserved(
        self, tmp_path: Path, hist_path: Path
    ) -> None:
        """Both 제품명(한글) and 제품명(영문) must be preserved in output."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(_MFDS_HEADERS)
        ws.append([
            "20260601", "한국수입사", "C1", "P1",
            "생녹용(냉동) 500g",          # 제품명(한글) — Korean name with weight
            "Frozen velvet antler 500g",  # 제품명(영문) — English name
            "생녹용(냉동)", "뉴질랜드", "뉴질랜드", "NZ Velvet Exports",
            "500", "100", "20280601", "뉴질랜드",
        ])
        p = tmp_path / "bilingual.xlsx"
        wb.save(str(p))

        from src.parsers.mfds_xlsx import parse
        rows = parse(p, xlsx_path_for_table=hist_path)
        assert rows[0]["product_ko"] == "생녹용(냉동) 500g"
        assert rows[0]["product_en"] == "Frozen velvet antler 500g"

    def test_non_numeric_weight_preserved_as_string(
        self, tmp_path: Path, hist_path: Path
    ) -> None:
        """Non-numeric weight values must not raise — just stored as string."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(_MFDS_HEADERS)
        ws.append([
            "20260601", "한국수입사", "D1", "Q1",
            "생녹용(냉동)", "Frozen velvet antler", "생녹용(냉동)",
            "뉴질랜드", "뉴질랜드", "NZ Velvet Exports",
            "N/A",    # non-numeric weight
            "50", "20280601", "뉴질랜드",
        ])
        p = tmp_path / "bad_weight.xlsx"
        wb.save(str(p))

        from src.parsers.mfds_xlsx import parse
        rows = parse(p, xlsx_path_for_table=hist_path)
        assert rows[0]["weight_kg"] == "N/A"
